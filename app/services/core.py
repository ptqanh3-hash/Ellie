from __future__ import annotations

import re
import subprocess
import tempfile
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.constants import (
    DEFAULT_DUE_SOON_DAYS,
    DEFAULT_SHEET_NAME,
    MANUAL_STATUSES,
    MASTER_CATEGORY_PIPELINE_STATUS,
    MASTER_CATEGORY_PRIORITY_STAGE,
    MASTER_CATEGORY_TASK_STATUS,
    MASTER_DEFAULTS,
    PIPELINE_STATUSES,
    PRIORITY_STAGES,
    TASK_PRIORITIES,
)
from app.database import DatabaseManager


class ValidationError(Exception):
    """Raised when app data does not satisfy business rules."""


@dataclass
class ImportReport:
    opportunities_created: int
    phases_created: int
    tasks_created: int
    skipped_rows: int
    source_path: str
    sheet_name: str


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if hasattr(value, "isoformat") and not isinstance(value, str):
        return value.isoformat()
    return str(value).strip()


def _parse_deadline(raw: Any) -> str | None:
    if raw in (None, ""):
        return None
    if isinstance(raw, datetime):
        return raw.date().isoformat()
    if hasattr(raw, "isoformat") and not isinstance(raw, str):
        return raw.isoformat()
    text = str(raw).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text


def compute_health_status(manual_status: str, deadline: str | None, due_soon_days: int = DEFAULT_DUE_SOON_DAYS) -> str:
    if manual_status == "Completed":
        return "Completed"
    if not deadline:
        return "No Deadline"
    try:
        due = datetime.fromisoformat(deadline).date()
    except ValueError:
        return "No Deadline"
    today = date.today()
    if due < today:
        return "Delayed"
    if due <= today + timedelta(days=due_soon_days):
        return "Due Soon"
    return "On Track"


def split_phase_from_pipeline(pipeline_title: str) -> tuple[str, str]:
    title = (pipeline_title or "").strip()
    if not title:
        return "", "General"
    match = re.search(r"(.+?)(?:/|\n)\s*(Phase\s*[\w\d-]+)\s*$", title, re.IGNORECASE)
    if match:
        return match.group(1).strip(), match.group(2).strip()
    return title, "General"


def _load_workbook_resilient(source: Path):
    try:
        return load_workbook(source, data_only=True)
    except (PermissionError, OSError):
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
            temp_path = Path(temp_file.name)
        command = [
            "powershell",
            "-NoProfile",
            "-Command",
            f"Copy-Item -LiteralPath '{source}' -Destination '{temp_path}' -Force",
        ]
        try:
            subprocess.run(command, check=True, capture_output=True, text=True)
            workbook = load_workbook(temp_path, data_only=True)
        finally:
            if temp_path.exists():
                temp_path.unlink()
        return workbook


class WorkspaceService:
    def __init__(self, db: DatabaseManager):
        self.db = db

    def ensure_workspace(self) -> None:
        self.db.initialize()


class MasterDataService:
    def __init__(self, db: DatabaseManager):
        self.db = db

    def list_values(self, category: str, include_inactive: bool = False) -> list[dict[str, Any]]:
        query = """
            SELECT id, category, value, sort_order, is_active
            FROM master_values
            WHERE category = ? AND archived_at IS NULL
        """
        params: list[Any] = [category]
        if not include_inactive:
            query += " AND is_active = 1"
        query += " ORDER BY sort_order, value COLLATE NOCASE"
        with self.db.connection() as conn:
            rows = conn.execute(query, params).fetchall()
        records = [dict(row) for row in rows]
        if records:
            return records
        defaults = MASTER_DEFAULTS.get(category, [])
        return [
            {"id": 0, "category": category, "value": value, "sort_order": index, "is_active": 1}
            for index, value in enumerate(defaults, start=1)
        ]

    def list_names(self, category: str, include_inactive: bool = False) -> list[str]:
        return [item["value"] for item in self.list_values(category, include_inactive=include_inactive)]

    def add_value(self, category: str, value: str) -> int:
        clean_value = _clean(value)
        if not clean_value:
            raise ValidationError("Value is required.")
        with self.db.connection() as conn:
            existing = conn.execute(
                """
                SELECT id, is_active
                FROM master_values
                WHERE category = ? AND value = ? AND archived_at IS NULL
                """,
                (category, clean_value),
            ).fetchone()
            if existing:
                if existing["is_active"]:
                    raise ValidationError("Value already exists.")
                conn.execute("UPDATE master_values SET is_active = 1 WHERE id = ?", (existing["id"],))
                return int(existing["id"])
            next_order = conn.execute(
                "SELECT COALESCE(MAX(sort_order), 0) + 1 AS next_order FROM master_values WHERE category = ? AND archived_at IS NULL",
                (category,),
            ).fetchone()["next_order"]
            cur = conn.execute(
                "INSERT INTO master_values (category, value, sort_order, is_active) VALUES (?, ?, ?, 1)",
                (category, clean_value, next_order),
            )
            return int(cur.lastrowid)

    def rename_value(self, category: str, value_id: int, new_value: str) -> None:
        clean_value = _clean(new_value)
        if not clean_value:
            raise ValidationError("Value is required.")
        with self.db.connection() as conn:
            existing = conn.execute(
                "SELECT id, value FROM master_values WHERE id = ? AND category = ? AND archived_at IS NULL",
                (value_id, category),
            ).fetchone()
            if not existing:
                raise ValidationError("Master data item not found.")
            duplicate = conn.execute(
                """
                SELECT id
                FROM master_values
                WHERE category = ? AND value = ? AND archived_at IS NULL AND id != ?
                """,
                (category, clean_value, value_id),
            ).fetchone()
            if duplicate:
                raise ValidationError("Value already exists.")
            conn.execute("UPDATE master_values SET value = ? WHERE id = ?", (clean_value, value_id))
            old_value = existing["value"]
            if category == MASTER_CATEGORY_TASK_STATUS:
                conn.execute("UPDATE tasks SET manual_status = ?, updated_at = CURRENT_TIMESTAMP WHERE manual_status = ?", (clean_value, old_value))
            elif category == MASTER_CATEGORY_PIPELINE_STATUS:
                conn.execute(
                    "UPDATE opportunities SET pipeline_status = ?, updated_at = CURRENT_TIMESTAMP WHERE pipeline_status = ?",
                    (clean_value, old_value),
                )
                conn.execute(
                    "UPDATE phases SET phase_status = ? WHERE COALESCE(phase_status, '') = COALESCE(?, '')",
                    (clean_value, old_value),
                )
            elif category == MASTER_CATEGORY_PRIORITY_STAGE:
                conn.execute(
                    "UPDATE opportunities SET priority_stage = ?, updated_at = CURRENT_TIMESTAMP WHERE priority_stage = ?",
                    (clean_value, old_value),
                )
                conn.execute("UPDATE phases SET name = ? WHERE name = ?", (clean_value, old_value))

    def deactivate_value(self, category: str, value_id: int) -> None:
        with self.db.connection() as conn:
            existing = conn.execute(
                "SELECT id FROM master_values WHERE id = ? AND category = ? AND archived_at IS NULL",
                (value_id, category),
            ).fetchone()
            if not existing:
                raise ValidationError("Master data item not found.")
            conn.execute("UPDATE master_values SET is_active = 0 WHERE id = ?", (value_id,))

    def task_statuses(self, include_inactive: bool = False) -> list[str]:
        return self.list_names(MASTER_CATEGORY_TASK_STATUS, include_inactive=include_inactive)

    def pipeline_statuses(self, include_inactive: bool = False) -> list[str]:
        return self.list_names(MASTER_CATEGORY_PIPELINE_STATUS, include_inactive=include_inactive)

    def priority_stages(self, include_inactive: bool = False) -> list[str]:
        return self.list_names(MASTER_CATEGORY_PRIORITY_STAGE, include_inactive=include_inactive)


class UserService:
    def __init__(self, db: DatabaseManager):
        self.db = db

    def ensure_user(self, display_name: str | None, conn=None) -> int | None:
        name = _clean(display_name)
        if not name:
            return None
        if conn is None:
            with self.db.connection() as own_conn:
                return self.ensure_user(name, conn=own_conn)
        row = conn.execute("SELECT id FROM users WHERE display_name = ?", (name,)).fetchone()
        if row:
            return row["id"]
        cur = conn.execute(
            "INSERT INTO users (display_name, color_seed) VALUES (?, ?)",
            (name, name[:1].upper()),
        )
        return int(cur.lastrowid)

    def list_users(self) -> list[dict[str, Any]]:
        with self.db.connection() as conn:
            rows = conn.execute(
                "SELECT id, display_name, email, is_active FROM users WHERE is_active = 1 ORDER BY display_name"
            ).fetchall()
        return [dict(row) for row in rows]

    def list_all_users(self) -> list[dict[str, Any]]:
        with self.db.connection() as conn:
            rows = conn.execute(
                "SELECT id, display_name, email, is_active FROM users ORDER BY is_active DESC, display_name"
            ).fetchall()
        return [dict(row) for row in rows]

    def rename_user(self, user_id: int, display_name: str) -> None:
        clean_name = _clean(display_name)
        if not clean_name:
            raise ValidationError("PIC name is required.")
        with self.db.connection() as conn:
            existing = conn.execute("SELECT id FROM users WHERE id = ?", (user_id,)).fetchone()
            if not existing:
                raise ValidationError("PIC not found.")
            duplicate = conn.execute(
                "SELECT id FROM users WHERE display_name = ? AND id != ?",
                (clean_name, user_id),
            ).fetchone()
            if duplicate:
                raise ValidationError("PIC already exists.")
            conn.execute("UPDATE users SET display_name = ? WHERE id = ?", (clean_name, user_id))

    def deactivate_user(self, user_id: int) -> None:
        with self.db.connection() as conn:
            existing = conn.execute("SELECT id FROM users WHERE id = ?", (user_id,)).fetchone()
            if not existing:
                raise ValidationError("PIC not found.")
            conn.execute("UPDATE users SET is_active = 0 WHERE id = ?", (user_id,))


class OpportunityService:
    def __init__(self, db: DatabaseManager):
        self.db = db
        self.users = UserService(db)
        self.master_data = MasterDataService(db)

    def create_opportunity(
        self,
        title: str,
        department_name: str,
        pipeline_status: str,
        priority_stage: str,
        detail: str = "",
        external_pic_name: str = "",
    ) -> int:
        title = _clean(title)
        department_name = _clean(department_name)
        pipeline_status = _clean(pipeline_status)
        priority_stage = _clean(priority_stage)
        detail = _clean(detail)
        external_pic_name = _clean(external_pic_name)

        if not title:
            raise ValidationError("Opportunity title is required.")
        valid_pipeline_statuses = self.master_data.pipeline_statuses(include_inactive=True) or PIPELINE_STATUSES
        valid_priority_stages = self.master_data.priority_stages(include_inactive=True) or PRIORITY_STAGES
        if not pipeline_status:
            raise ValidationError("Pipeline status is required.")
        if pipeline_status not in valid_pipeline_statuses:
            raise ValidationError("Pipeline status is invalid.")
        if not priority_stage:
            raise ValidationError("Priority stage is required.")
        if priority_stage not in valid_priority_stages:
            raise ValidationError("Priority stage is invalid.")

        with self.db.connection() as conn:
            external_pic_id = None
            if external_pic_name:
                row = conn.execute(
                    "SELECT id FROM contacts WHERE name = ? AND COALESCE(department_name, '') = COALESCE(?, '')",
                    (external_pic_name, department_name or None),
                ).fetchone()
                if row:
                    external_pic_id = row["id"]
                else:
                    cur = conn.execute(
                        "INSERT INTO contacts (name, department_name) VALUES (?, ?)",
                        (external_pic_name, department_name or None),
                    )
                    external_pic_id = int(cur.lastrowid)

            cur = conn.execute(
                """
                INSERT INTO opportunities (
                    title, department_name, external_pic_id, pipeline_status, priority_stage, detail
                ) VALUES (?, ?, ?, ?, ?, ?)
                """,
                (title, department_name or None, external_pic_id, pipeline_status, priority_stage, detail or None),
            )
            opportunity_id = int(cur.lastrowid)
            conn.execute(
                "INSERT INTO phases (opportunity_id, name, phase_status, order_index) VALUES (?, ?, ?, ?)",
                (opportunity_id, priority_stage or "General", pipeline_status or None, 1),
            )
            return opportunity_id

    def update_opportunity(
        self,
        opportunity_id: int,
        title: str,
        department_name: str,
        pipeline_status: str,
        priority_stage: str,
        detail: str = "",
        external_pic_name: str = "",
    ) -> None:
        title = _clean(title)
        department_name = _clean(department_name)
        pipeline_status = _clean(pipeline_status)
        priority_stage = _clean(priority_stage)
        detail = _clean(detail)
        external_pic_name = _clean(external_pic_name)
        valid_pipeline_statuses = self.master_data.pipeline_statuses(include_inactive=True) or PIPELINE_STATUSES
        valid_priority_stages = self.master_data.priority_stages(include_inactive=True) or PRIORITY_STAGES

        if not title:
            raise ValidationError("Opportunity title is required.")
        if pipeline_status not in valid_pipeline_statuses:
            raise ValidationError("Pipeline status is invalid.")
        if priority_stage not in valid_priority_stages:
            raise ValidationError("Priority stage is invalid.")

        with self.db.connection() as conn:
            existing = conn.execute(
                "SELECT id FROM opportunities WHERE id = ? AND archived_at IS NULL",
                (opportunity_id,),
            ).fetchone()
            if not existing:
                raise ValidationError("Opportunity not found.")
            external_pic_id = None
            if external_pic_name:
                row = conn.execute(
                    "SELECT id FROM contacts WHERE name = ? AND COALESCE(department_name, '') = COALESCE(?, '')",
                    (external_pic_name, department_name or None),
                ).fetchone()
                if row:
                    external_pic_id = row["id"]
                else:
                    cur = conn.execute(
                        "INSERT INTO contacts (name, department_name) VALUES (?, ?)",
                        (external_pic_name, department_name or None),
                    )
                    external_pic_id = int(cur.lastrowid)
            conn.execute(
                """
                UPDATE opportunities
                SET title = ?, department_name = ?, external_pic_id = ?, pipeline_status = ?,
                    priority_stage = ?, detail = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (title, department_name or None, external_pic_id, pipeline_status, priority_stage, detail or None, opportunity_id),
            )

    def archive_opportunity(self, opportunity_id: int) -> None:
        with self.db.connection() as conn:
            existing = conn.execute(
                "SELECT id FROM opportunities WHERE id = ? AND archived_at IS NULL",
                (opportunity_id,),
            ).fetchone()
            if not existing:
                raise ValidationError("Opportunity not found.")
            timestamp = datetime.utcnow().isoformat()
            conn.execute("UPDATE opportunities SET archived_at = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?", (timestamp, opportunity_id))
            conn.execute("UPDATE phases SET archived_at = ? WHERE opportunity_id = ? AND archived_at IS NULL", (timestamp, opportunity_id))
            conn.execute("UPDATE tasks SET archived_at = ?, updated_at = CURRENT_TIMESTAMP WHERE opportunity_id = ? AND archived_at IS NULL", (timestamp, opportunity_id))

    def create_stage(self, opportunity_id: int, stage_name: str, stage_status: str, description: str = "") -> int:
        stage_name = _clean(stage_name)
        stage_status = _clean(stage_status)
        valid_stage_statuses = self.master_data.pipeline_statuses(include_inactive=True) or PIPELINE_STATUSES
        if not stage_name:
            raise ValidationError("Stage name is required.")
        if not stage_status:
            raise ValidationError("Stage status is required.")
        if stage_status not in valid_stage_statuses:
            raise ValidationError("Stage status is invalid.")
        with self.db.connection() as conn:
            duplicate = conn.execute(
                """
                SELECT id
                FROM phases
                WHERE opportunity_id = ? AND name = ? AND COALESCE(phase_status, '') = COALESCE(?, '') AND archived_at IS NULL
                """,
                (opportunity_id, stage_name, stage_status or None),
            ).fetchone()
            if duplicate:
                raise ValidationError("Stage + status already exists.")
            row = conn.execute(
                "SELECT COALESCE(MAX(order_index), 0) + 1 AS next_order FROM phases WHERE opportunity_id = ?",
                (opportunity_id,),
            ).fetchone()
            cur = conn.execute(
                """
                INSERT INTO phases (opportunity_id, name, description, phase_status, order_index)
                VALUES (?, ?, ?, ?, ?)
                """,
                (opportunity_id, stage_name, _clean(description) or None, stage_status or None, row["next_order"]),
            )
            return int(cur.lastrowid)

    def create_phase(self, opportunity_id: int, name: str, description: str = "") -> int:
        return self.create_stage(opportunity_id, name, PIPELINE_STATUSES[0], description)

    def update_stage(self, stage_id: int, stage_name: str, stage_status: str, description: str = "") -> None:
        stage_name = _clean(stage_name)
        stage_status = _clean(stage_status)
        valid_stage_statuses = self.master_data.pipeline_statuses(include_inactive=True) or PIPELINE_STATUSES
        if not stage_name:
            raise ValidationError("Stage name is required.")
        if not stage_status:
            raise ValidationError("Stage status is required.")
        if stage_status not in valid_stage_statuses:
            raise ValidationError("Stage status is invalid.")
        with self.db.connection() as conn:
            existing = conn.execute(
                "SELECT id, opportunity_id FROM phases WHERE id = ? AND archived_at IS NULL",
                (stage_id,),
            ).fetchone()
            if not existing:
                raise ValidationError("Stage not found.")
            duplicate = conn.execute(
                """
                SELECT id
                FROM phases
                WHERE opportunity_id = ? AND name = ? AND COALESCE(phase_status, '') = COALESCE(?, '')
                    AND archived_at IS NULL AND id != ?
                """,
                (existing["opportunity_id"], stage_name, stage_status or None, stage_id),
            ).fetchone()
            if duplicate:
                raise ValidationError("Stage + status already exists.")
            conn.execute(
                """
                UPDATE phases
                SET name = ?, phase_status = ?, description = ?
                WHERE id = ?
                """,
                (stage_name, stage_status or None, _clean(description) or None, stage_id),
            )

    def archive_stage(self, stage_id: int) -> None:
        with self.db.connection() as conn:
            existing = conn.execute(
                "SELECT id FROM phases WHERE id = ? AND archived_at IS NULL",
                (stage_id,),
            ).fetchone()
            if not existing:
                raise ValidationError("Stage not found.")
            timestamp = datetime.utcnow().isoformat()
            conn.execute("UPDATE phases SET archived_at = ? WHERE id = ?", (timestamp, stage_id))
            conn.execute("UPDATE tasks SET archived_at = ?, updated_at = CURRENT_TIMESTAMP WHERE phase_id = ? AND archived_at IS NULL", (timestamp, stage_id))

    def list_opportunities(self) -> list[dict[str, Any]]:
        with self.db.connection() as conn:
            rows = conn.execute(
                """
                SELECT
                    o.id,
                    o.title,
                    o.department_name,
                    c.name AS external_pic_name,
                    o.pipeline_status,
                    o.priority_stage,
                    COUNT(DISTINCT t.id) AS task_count,
                    SUM(CASE WHEN t.manual_status != 'Completed' THEN 1 ELSE 0 END) AS open_task_count
                FROM opportunities o
                LEFT JOIN contacts c ON c.id = o.external_pic_id
                LEFT JOIN tasks t ON t.opportunity_id = o.id AND t.archived_at IS NULL
                WHERE o.archived_at IS NULL
                GROUP BY o.id
                ORDER BY o.updated_at DESC, o.id DESC
                """
            ).fetchall()
        return [dict(row) for row in rows]

    def get_opportunity_detail(self, opportunity_id: int) -> dict[str, Any]:
        with self.db.connection() as conn:
            opportunity = conn.execute(
                """
                SELECT o.*, c.name AS external_pic_name
                FROM opportunities o
                LEFT JOIN contacts c ON c.id = o.external_pic_id
                WHERE o.id = ?
                """,
                (opportunity_id,),
            ).fetchone()
            if not opportunity:
                raise ValidationError("Opportunity not found.")
            phases = conn.execute(
                "SELECT * FROM phases WHERE opportunity_id = ? AND archived_at IS NULL ORDER BY order_index, id",
                (opportunity_id,),
            ).fetchall()
            tasks = conn.execute(
                """
                SELECT
                    t.*,
                    p.name AS stage_name,
                    p.phase_status AS stage_status,
                    owner.display_name AS owner_name,
                    pic.display_name AS pic_name
                FROM tasks t
                JOIN phases p ON p.id = t.phase_id
                LEFT JOIN users owner ON owner.id = t.owner_user_id
                LEFT JOIN users pic ON pic.id = t.pic_user_id
                WHERE t.opportunity_id = ? AND t.archived_at IS NULL
                ORDER BY t.deadline IS NULL, t.deadline, t.sort_order, t.id
                """,
                (opportunity_id,),
            ).fetchall()
        task_items = []
        for task in tasks:
            item = dict(task)
            item["health_status"] = compute_health_status(item["manual_status"], item["deadline"])
            item["stage_label"] = f"{item.get('stage_name') or '-'} | {item.get('stage_status') or '-'}"
            task_items.append(item)
        stage_items = []
        for row in phases:
            item = dict(row)
            item["stage_name"] = item.get("name")
            item["stage_status"] = item.get("phase_status") or opportunity["pipeline_status"]
            item["task_count"] = sum(1 for task in task_items if task["phase_id"] == item["id"])
            stage_items.append(item)
        return {
            "opportunity": dict(opportunity),
            "stages": stage_items,
            "phases": [dict(row) for row in phases],
            "tasks": task_items,
        }


class TaskService:
    def __init__(self, db: DatabaseManager):
        self.db = db
        self.users = UserService(db)
        self.master_data = MasterDataService(db)

    def create_task(
        self,
        opportunity_id: int,
        phase_id: int,
        title: str,
        owner_name: str = "",
        pic_name: str = "",
        manual_status: str = "Not Started",
        deadline: str | None = None,
        next_action: str = "",
        description: str = "",
        priority: str = "Normal",
        latest_update_summary: str = "",
    ) -> int:
        title = _clean(title)
        manual_status = _clean(manual_status) or "Not Started"
        priority = _clean(priority) or "Normal"
        valid_statuses = self.master_data.task_statuses(include_inactive=True) or MANUAL_STATUSES
        if not title:
            raise ValidationError("Task title is required.")
        if manual_status not in valid_statuses:
            raise ValidationError("Task status is invalid.")
        if priority not in TASK_PRIORITIES:
            raise ValidationError("Task priority is invalid.")

        with self.db.connection() as conn:
            owner_id = self.users.ensure_user(owner_name, conn=conn)
            pic_id = self.users.ensure_user(pic_name, conn=conn)
            row = conn.execute(
                "SELECT COALESCE(MAX(sort_order), 0) + 1 AS next_order FROM tasks WHERE phase_id = ?",
                (phase_id,),
            ).fetchone()
            completed_at = datetime.utcnow().isoformat() if manual_status == "Completed" else None
            cur = conn.execute(
                """
                INSERT INTO tasks (
                    opportunity_id, phase_id, title, description, owner_user_id, pic_user_id,
                    manual_status, priority, deadline, completed_at, next_action,
                    latest_update_summary, sort_order
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    opportunity_id,
                    phase_id,
                    title,
                    _clean(description) or None,
                    owner_id,
                    pic_id,
                    manual_status,
                    priority,
                    _parse_deadline(deadline),
                    completed_at,
                    _clean(next_action) or None,
                    _clean(latest_update_summary) or None,
                    row["next_order"],
                ),
            )
            task_id = int(cur.lastrowid)
            if latest_update_summary:
                conn.execute(
                    "INSERT INTO task_updates (task_id, author_user_id, entry_type, content) VALUES (?, ?, ?, ?)",
                    (task_id, owner_id or pic_id, "note", latest_update_summary),
                )
            return task_id

    def get_task(self, task_id: int) -> dict[str, Any]:
        with self.db.connection() as conn:
            row = conn.execute(
                """
                SELECT
                    t.*,
                    o.title AS opportunity_title,
                    p.name AS phase_name,
                    p.name AS stage_name,
                    p.phase_status AS stage_status,
                    owner.display_name AS owner_name,
                    pic.display_name AS pic_name
                FROM tasks t
                JOIN opportunities o ON o.id = t.opportunity_id
                JOIN phases p ON p.id = t.phase_id
                LEFT JOIN users owner ON owner.id = t.owner_user_id
                LEFT JOIN users pic ON pic.id = t.pic_user_id
                WHERE t.id = ? AND t.archived_at IS NULL
                """,
                (task_id,),
            ).fetchone()
        if not row:
            raise ValidationError("Task not found.")
        task = dict(row)
        task["health_status"] = compute_health_status(task["manual_status"], task["deadline"])
        task["stage_label"] = f"{task.get('stage_name') or '-'} | {task.get('stage_status') or '-'}"
        return task

    def update_task(
        self,
        task_id: int,
        phase_id: int,
        title: str,
        owner_name: str = "",
        pic_name: str = "",
        manual_status: str = "Not Started",
        deadline: str | None = None,
        next_action: str = "",
        description: str = "",
        priority: str = "Normal",
        latest_update_summary: str = "",
    ) -> None:
        title = _clean(title)
        manual_status = _clean(manual_status) or "Not Started"
        priority = _clean(priority) or "Normal"
        valid_statuses = self.master_data.task_statuses(include_inactive=True) or MANUAL_STATUSES
        if not title:
            raise ValidationError("Task title is required.")
        if manual_status not in valid_statuses:
            raise ValidationError("Task status is invalid.")
        if priority not in TASK_PRIORITIES:
            raise ValidationError("Task priority is invalid.")

        with self.db.connection() as conn:
            existing = conn.execute(
                "SELECT id, opportunity_id FROM tasks WHERE id = ? AND archived_at IS NULL",
                (task_id,),
            ).fetchone()
            if not existing:
                raise ValidationError("Task not found.")
            phase = conn.execute(
                "SELECT id, opportunity_id FROM phases WHERE id = ? AND archived_at IS NULL",
                (phase_id,),
            ).fetchone()
            if not phase or phase["opportunity_id"] != existing["opportunity_id"]:
                raise ValidationError("Selected phase is invalid.")
            owner_id = self.users.ensure_user(owner_name, conn=conn)
            pic_id = self.users.ensure_user(pic_name, conn=conn)
            completed_at = datetime.utcnow().isoformat() if manual_status == "Completed" else None
            conn.execute(
                """
                UPDATE tasks
                SET phase_id = ?, title = ?, description = ?, owner_user_id = ?, pic_user_id = ?,
                    manual_status = ?, priority = ?, deadline = ?, completed_at = ?, next_action = ?,
                    latest_update_summary = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (
                    phase_id,
                    title,
                    _clean(description) or None,
                    owner_id,
                    pic_id,
                    manual_status,
                    priority,
                    _parse_deadline(deadline),
                    completed_at,
                    _clean(next_action) or None,
                    _clean(latest_update_summary) or None,
                    task_id,
                ),
            )
            if latest_update_summary:
                conn.execute(
                    "INSERT INTO task_updates (task_id, author_user_id, entry_type, content) VALUES (?, ?, ?, ?)",
                    (task_id, owner_id or pic_id, "note", latest_update_summary),
                )

    def archive_task(self, task_id: int) -> None:
        with self.db.connection() as conn:
            existing = conn.execute(
                "SELECT id FROM tasks WHERE id = ? AND archived_at IS NULL",
                (task_id,),
            ).fetchone()
            if not existing:
                raise ValidationError("Task not found.")
            conn.execute(
                "UPDATE tasks SET archived_at = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                (datetime.utcnow().isoformat(), task_id),
            )

    def update_task_status(self, task_id: int, manual_status: str) -> None:
        manual_status = _clean(manual_status)
        valid_statuses = self.master_data.task_statuses(include_inactive=True) or MANUAL_STATUSES
        if manual_status not in valid_statuses:
            raise ValidationError("Task status is invalid.")
        with self.db.connection() as conn:
            existing = conn.execute("SELECT id FROM tasks WHERE id = ? AND archived_at IS NULL", (task_id,)).fetchone()
            if not existing:
                raise ValidationError("Task not found.")
            completed_at = datetime.utcnow().isoformat() if manual_status == "Completed" else None
            conn.execute(
                """
                UPDATE tasks
                SET manual_status = ?, completed_at = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (manual_status, completed_at, task_id),
            )
            conn.execute(
                "INSERT INTO task_updates (task_id, entry_type, content) VALUES (?, ?, ?)",
                (task_id, "status_change", f"Status changed to {manual_status}"),
            )

    def list_tasks(self) -> list[dict[str, Any]]:
        with self.db.connection() as conn:
            rows = conn.execute(
                """
                SELECT
                    t.*,
                    o.title AS opportunity_title,
                    p.name AS phase_name,
                    p.name AS stage_name,
                    p.phase_status AS stage_status,
                    owner.display_name AS owner_name,
                    pic.display_name AS pic_name
                FROM tasks t
                JOIN opportunities o ON o.id = t.opportunity_id
                JOIN phases p ON p.id = t.phase_id
                LEFT JOIN users owner ON owner.id = t.owner_user_id
                LEFT JOIN users pic ON pic.id = t.pic_user_id
                WHERE t.archived_at IS NULL
                ORDER BY t.deadline IS NULL, t.deadline, t.id
                """
            ).fetchall()
        items = []
        for row in rows:
            item = dict(row)
            item["health_status"] = compute_health_status(item["manual_status"], item["deadline"])
            item["stage_label"] = f"{item.get('stage_name') or '-'} | {item.get('stage_status') or '-'}"
            items.append(item)
        return items

    def board_columns(self) -> dict[str, list[dict[str, Any]]]:
        statuses = self.master_data.task_statuses() or MANUAL_STATUSES
        columns = {status: [] for status in statuses}
        for task in self.list_tasks():
            columns.setdefault(task["manual_status"], []).append(task)
        return columns


class DashboardService:
    def __init__(self, db: DatabaseManager):
        self.db = db
        self.tasks = TaskService(db)

    def metrics(self) -> dict[str, Any]:
        tasks = self.tasks.list_tasks()
        opportunities = OpportunityService(self.db).list_opportunities()
        overdue = [task for task in tasks if task["health_status"] == "Delayed"]
        due_soon = [task for task in tasks if task["health_status"] == "Due Soon"]
        recent = sorted(tasks, key=lambda item: (item["updated_at"], item["id"]), reverse=True)[:5]
        return {
            "opportunity_count": len(opportunities),
            "task_count": len(tasks),
            "overdue_count": len(overdue),
            "due_soon_count": len(due_soon),
            "recent_tasks": recent,
            "overdue_tasks": overdue[:8],
            "due_soon_tasks": due_soon[:8],
        }


class ExcelImportService:
    def __init__(self, db: DatabaseManager):
        self.db = db
        self.workspace = WorkspaceService(db)
        self.users = UserService(db)
        self.master_data = MasterDataService(db)

    def import_workbook(self, workbook_path: str | Path, sheet_name: str = DEFAULT_SHEET_NAME) -> ImportReport:
        source = Path(workbook_path)
        if not source.exists():
            raise ValidationError(f"Workbook not found: {source}")

        self.workspace.ensure_workspace()
        workbook = _load_workbook_resilient(source)
        if sheet_name not in workbook.sheetnames:
            raise ValidationError(f"Worksheet '{sheet_name}' not found.")
        sheet = workbook[sheet_name]

        opportunities_created = 0
        phases_created = 0
        tasks_created = 0
        skipped_rows = 0
        current_context = {key: "" for key in ("B", "C", "D", "E", "F", "G")}
        opportunity_cache: dict[tuple[str, ...], int] = {}
        phase_cache: dict[tuple[int, str], int] = {}
        contact_cache: dict[tuple[str, str], int] = {}
        pipeline_statuses = self.master_data.pipeline_statuses(include_inactive=True) or PIPELINE_STATUSES
        priority_stages = self.master_data.priority_stages(include_inactive=True) or PRIORITY_STAGES
        task_statuses = self.master_data.task_statuses(include_inactive=True) or MANUAL_STATUSES

        with self.db.connection() as conn:
            for row_idx in range(6, sheet.max_row + 1):
                row = {
                    "B": _clean(sheet[f"B{row_idx}"].value),
                    "C": _clean(sheet[f"C{row_idx}"].value),
                    "D": _clean(sheet[f"D{row_idx}"].value),
                    "E": _clean(sheet[f"E{row_idx}"].value),
                    "F": _clean(sheet[f"F{row_idx}"].value),
                    "G": _clean(sheet[f"G{row_idx}"].value),
                    "H": _clean(sheet[f"H{row_idx}"].value),
                    "I": _clean(sheet[f"I{row_idx}"].value),
                    "J": _clean(sheet[f"J{row_idx}"].value),
                    "K": sheet[f"K{row_idx}"].value,
                    "L": _clean(sheet[f"L{row_idx}"].value),
                    "N": _clean(sheet[f"N{row_idx}"].value),
                    "P": _clean(sheet[f"P{row_idx}"].value),
                }
                if not any(row.values()):
                    continue

                for key in ("B", "C", "D", "E", "F", "G"):
                    if row[key]:
                        current_context[key] = row[key]

                if not current_context["D"]:
                    skipped_rows += 1
                    continue

                title, phase_name = split_phase_from_pipeline(current_context["D"])
                pipeline_status = current_context["F"] if current_context["F"] in pipeline_statuses else (current_context["F"] or pipeline_statuses[0])
                priority_stage = current_context["G"] if current_context["G"] in priority_stages else (current_context["G"] or priority_stages[0])

                external_pic_id = None
                contact_key = (current_context["C"], current_context["B"])
                if current_context["C"]:
                    if contact_key in contact_cache:
                        external_pic_id = contact_cache[contact_key]
                    else:
                        existing = conn.execute(
                            "SELECT id FROM contacts WHERE name = ? AND COALESCE(department_name, '') = COALESCE(?, '')",
                            (current_context["C"], current_context["B"] or None),
                        ).fetchone()
                        if existing:
                            external_pic_id = existing["id"]
                        else:
                            cur = conn.execute(
                                "INSERT INTO contacts (name, department_name) VALUES (?, ?)",
                                (current_context["C"], current_context["B"] or None),
                            )
                            external_pic_id = int(cur.lastrowid)
                        contact_cache[contact_key] = external_pic_id

                opportunity_key = (
                    title,
                    current_context["B"],
                    str(external_pic_id or ""),
                    pipeline_status,
                    priority_stage,
                    current_context["E"],
                )
                if opportunity_key in opportunity_cache:
                    opportunity_id = opportunity_cache[opportunity_key]
                else:
                    cur = conn.execute(
                        """
                        INSERT INTO opportunities (
                            title, department_name, external_pic_id, pipeline_status, priority_stage, detail, source_row_ref
                        ) VALUES (?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            title,
                            current_context["B"] or None,
                            external_pic_id,
                            pipeline_status,
                            priority_stage,
                            current_context["E"] or None,
                            f"row:{row_idx}",
                        ),
                    )
                    opportunity_id = int(cur.lastrowid)
                    opportunity_cache[opportunity_key] = opportunity_id
                    opportunities_created += 1

                phase_key = (opportunity_id, phase_name)
                if phase_key in phase_cache:
                    phase_id = phase_cache[phase_key]
                else:
                    existing_phase = conn.execute(
                        "SELECT id FROM phases WHERE opportunity_id = ? AND name = ?",
                        (opportunity_id, phase_name),
                    ).fetchone()
                    if existing_phase:
                        phase_id = existing_phase["id"]
                    else:
                        next_order = conn.execute(
                            "SELECT COALESCE(MAX(order_index), 0) + 1 AS next_order FROM phases WHERE opportunity_id = ?",
                            (opportunity_id,),
                        ).fetchone()["next_order"]
                        cur = conn.execute(
                            "INSERT INTO phases (opportunity_id, name, phase_status, order_index) VALUES (?, ?, ?, ?)",
                            (opportunity_id, phase_name, pipeline_status or None, next_order),
                        )
                        phase_id = int(cur.lastrowid)
                        phases_created += 1
                    phase_cache[phase_key] = phase_id

                if not row["H"]:
                    continue

                owner_id = self.users.ensure_user(row["I"], conn=conn)
                pic_id = self.users.ensure_user(row["J"], conn=conn)
                manual_status = row["L"] if row["L"] in task_statuses else task_statuses[0]
                completed_at = datetime.utcnow().isoformat() if manual_status == "Completed" else None
                deadline = _parse_deadline(row["K"])
                cur = conn.execute(
                    """
                    INSERT INTO tasks (
                        opportunity_id, phase_id, title, owner_user_id, pic_user_id, manual_status,
                        priority, deadline, completed_at, next_action, latest_update_summary, source_row_ref
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        opportunity_id,
                        phase_id,
                        row["H"],
                        owner_id,
                        pic_id,
                        manual_status,
                        "Normal",
                        deadline,
                        completed_at,
                        row["P"] or None,
                        row["N"] or None,
                        f"row:{row_idx}",
                    ),
                )
                task_id = int(cur.lastrowid)
                tasks_created += 1
                if row["N"]:
                    conn.execute(
                        "INSERT INTO task_updates (task_id, author_user_id, entry_type, content) VALUES (?, ?, ?, ?)",
                        (task_id, owner_id or pic_id, "note", row["N"]),
                    )

            conn.execute(
                """
                INSERT INTO imports (
                    source_path, sheet_name, opportunities_created, phases_created, tasks_created, skipped_rows
                ) VALUES (?, ?, ?, ?, ?, ?)
                """,
                (str(source), sheet_name, opportunities_created, phases_created, tasks_created, skipped_rows),
            )

        return ImportReport(
            opportunities_created=opportunities_created,
            phases_created=phases_created,
            tasks_created=tasks_created,
            skipped_rows=skipped_rows,
            source_path=str(source),
            sheet_name=sheet_name,
        )
